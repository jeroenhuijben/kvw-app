#!/usr/bin/env python3
"""Convert the KVW Excel timetable to browser-ready JavaScript data."""

import argparse
import json
from collections import defaultdict
from datetime import time
from pathlib import Path

from openpyxl import load_workbook


ROOMS = [
    "Gymzaal Laagstraat",
    "Gymzaal Wilgenstraat",
    "Aula",
    "Kleuteraula",
    "Binnenplaats",
    "Sportveld voor",
    "Sportveld achter",
    "Extern 1",
    "Extern 2",
    "Extern 3",
]

DAY_DATES = {
    "Maandag": "17 aug",
    "Dinsdag": "18 aug",
    "Woensdag": "19 aug",
    "Donderdag": "20 aug",
    "Vrijdag": "21 aug",
}

LEGEND = [
    {"type": "kids", "label": "Kleuters"},
    {"type": "kids-pupils", "label": "Kleuters + Pupillen"},
    {"type": "pupils", "label": "Pupillen"},
    {"type": "pupils-youth", "label": "Pupillen + Jongeren"},
    {"type": "youth", "label": "Jongeren"},
    {"type": "youth-older", "label": "Jongeren + Ouderen"},
    {"type": "older", "label": "Ouderen"},
    {"type": "general", "label": "Alle groepen"},
]

FILL_TYPES = {
    ("rgb", "FFFFC000"): "kids",
    ("rgb", "FFFFFF00"): "kids-pupils",
    ("rgb", "FF00B0F0"): "pupils",
    ("theme", 7): "pupils",
    ("rgb", "FF0070C0"): "pupils-youth",
    ("rgb", "FF92D050"): "youth",
    ("rgb", "FF00B050"): "youth-older",
    ("theme", 9): "youth-older",
    ("rgb", "FFFF0000"): "older",
    ("rgb", "FFF1CEEE"): "general",
}


def parse_args():
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, default=Path("timetable-data.js"))
    return parser.parse_args()


def fill_type(cell):
    color = cell.fill.fgColor
    value = color.rgb if color.type == "rgb" else color.theme
    return FILL_TYPES.get((color.type, value), "default")


def normalized_activity(value):
    return " · ".join(part.strip() for part in str(value).splitlines() if part.strip())


def minutes_for_row(row):
    return (9 * 60) + 15 + ((row - 2) * 15)


def format_time(minutes):
    return f"{minutes // 60:02d}:{minutes % 60:02d}"


def location_label(column, span):
    labels = ROOMS[column : column + span]
    if len(labels) == len(ROOMS):
        return "Alle locaties"
    return " + ".join(labels)


def merged_ranges_by_origin(sheet):
    return {
        (merged.min_row, merged.min_col): merged
        for merged in sheet.merged_cells.ranges
        if merged.min_row <= 29 and merged.max_row >= 2
    }


def sheet_to_day(sheet):
    merged_ranges = merged_ranges_by_origin(sheet)
    grouped_rows = defaultdict(list)

    for row in range(2, 30):
        for column in range(2, 12):
            cell = sheet.cell(row=row, column=column)
            if cell.value is None:
                continue

            merged = merged_ranges.get((row, column))
            row_span = merged.max_row - merged.min_row + 1 if merged else 1
            column_span = merged.max_col - merged.min_col + 1 if merged else 1
            column_index = column - 2
            start = minutes_for_row(row)
            end = start + (row_span * 15)
            time_range = f"{format_time(start)} - {format_time(end)}"

            grouped_rows[time_range].append(
                {
                    "column": column_index,
                    "columnSpan": column_span,
                    "room": location_label(column_index, column_span),
                    "activity": normalized_activity(cell.value),
                    "type": fill_type(cell),
                }
            )

    rows = [
        {"time": time_range, "rooms": rooms}
        for time_range, rooms in sorted(
            grouped_rows.items(),
            key=lambda item: int(item[0][:2]) * 60 + int(item[0][3:5]),
        )
    ]
    return {"label": sheet.title, "date": DAY_DATES[sheet.title], "rows": rows}


def tuesday_to_verland():
    return {
        "label": "Dinsdag",
        "date": DAY_DATES["Dinsdag"],
        "rows": [
            {
                "time": "09:15 - 16:00",
                "rooms": [
                    {
                        "column": 0,
                        "columnSpan": len(ROOMS),
                        "room": "Toverland",
                        "activity": "Alle groepen · Toverland",
                        "type": "general",
                    }
                ],
            }
        ],
    }


def main():
    args = parse_args()
    workbook = load_workbook(args.input, data_only=False)
    days_by_name = {sheet.title: sheet_to_day(sheet) for sheet in workbook.worksheets}
    days = []
    for day_name in DAY_DATES:
        if day_name in days_by_name:
            days.append(days_by_name[day_name])
        elif day_name == "Dinsdag":
            days.append(tuesday_to_verland())
        else:
            raise ValueError(f"Werkblad ontbreekt: {day_name}")

    data = {
        "source": args.input.name,
        "start": "09:15",
        "end": "16:15",
        "rooms": ROOMS,
        "legend": LEGEND,
        "days": days,
    }
    output = "// Generated from Timetable.xlsx.\nwindow.KVW_TIMETABLE_DATA = "
    output += json.dumps(data, ensure_ascii=False, indent=2)
    output += ";\n"
    args.output.write_text(output, encoding="utf-8")


if __name__ == "__main__":
    main()
